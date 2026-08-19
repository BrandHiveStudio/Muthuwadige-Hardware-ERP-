#!/usr/bin/env python3
"""
Template-Based Excel Backup Generator
Loads reference template, populates with live data, preserves all formatting
"""

import json
import sys
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def populate_worksheet(ws, data_rows, headers):
    """Populate worksheet with data while preserving template structure and formatting"""
    if not data_rows:
        return

    # Find where to start inserting data (after header rows)
    # Most sheets have headers in row 1
    start_row = 2

    # Insert data rows
    for row_idx, row_data in enumerate(data_rows, start=start_row):
        for col_idx, (header, value) in enumerate(zip(headers, row_data), start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value

            # Preserve cell formatting from template if it exists
            if row_idx > 2 and ws.cell(row=row_idx - 1, column=col_idx):
                prev_cell = ws.cell(row=row_idx - 1, column=col_idx)
                if hasattr(prev_cell, 'font') and prev_cell.font:
                    cell.font = prev_cell.font.copy()
                if hasattr(prev_cell, 'fill') and prev_cell.fill:
                    cell.fill = prev_cell.fill.copy()
                if hasattr(prev_cell, 'alignment') and prev_cell.alignment:
                    cell.alignment = prev_cell.alignment.copy()
                if hasattr(prev_cell, 'border') and prev_cell.border:
                    cell.border = prev_cell.border.copy()
                if hasattr(prev_cell, 'number_format'):
                    cell.number_format = prev_cell.number_format

def main():
    if len(sys.argv) < 4:
        print("Usage: python backup-template-generator.py <template_path> <output_path> <data_json>")
        sys.exit(1)

    template_path = sys.argv[1]
    output_path = sys.argv[2]
    data_json = sys.argv[3]

    try:
        # Load template
        if not Path(template_path).exists():
            print(f"Template not found: {template_path}")
            sys.exit(1)

        wb = load_workbook(template_path)
        print(f"Template loaded: {template_path}")
        print(f"Sheets: {wb.sheetnames}")

        # Parse data
        data = json.loads(data_json)
        print(f"Data received: {len(data)} sheet specifications")

        # Populate each sheet
        for sheet_name, sheet_data in data.items():
            if sheet_name not in wb.sheetnames:
                print(f"Warning: Sheet '{sheet_name}' not in template, skipping")
                continue

            ws = wb[sheet_name]
            headers = sheet_data.get('headers', [])
            rows = sheet_data.get('rows', [])

            if headers and rows:
                # Clear existing data rows (preserve header and structure)
                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        if cell.value is not None:
                            cell.value = None

                # Populate with new data
                populate_worksheet(ws, rows, headers)
                print(f"  Populated '{sheet_name}': {len(rows)} rows")
            else:
                print(f"  Skipped '{sheet_name}': No data provided")

        # Save file
        wb.save(output_path)
        print(f"Workbook saved: {output_path}")
        print("Success")
        sys.exit(0)

    except Exception as e:
        print(f"Error: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == '__main__':
    main()
